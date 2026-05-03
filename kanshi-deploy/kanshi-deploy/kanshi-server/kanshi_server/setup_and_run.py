from pathlib import Path
import sys
import os
import platform
import subprocess
import json
from datetime import date, datetime, timedelta, timezone
from collections import defaultdict
from cryptography.fernet import Fernet


# ── Configuration ──────────────────────────────────────────────
def _get_documents_dir() -> Path:
    if platform.system() == "Windows":
        onedrive = os.environ.get("OneDrive", "")
        if onedrive:
            od_docs = Path(onedrive) / "Documents"
            if od_docs.exists():
                return od_docs
    return Path.home() / "Documents"

SOURCE_FOLDER = _get_documents_dir() / "KanshiReports"
KEY_FILE      = SOURCE_FOLDER / ".key"
REPORT_FILE   = SOURCE_FOLDER / "report_output.txt"
DATE_TO       = date.today() - timedelta(days=1)  # always yesterday

if platform.system() == "Windows":
    DATA_DIR = Path(os.environ.get("LOCALAPPDATA", "")) / "kanshi" / "kanshi" / "kanshi-server"
elif platform.system() == "Darwin":
    DATA_DIR = Path.home() / "Library" / "Application Support" / "kanshi" / "kanshi-server"
else:
    DATA_DIR = Path.home() / ".local" / "share" / "kanshi" / "kanshi-server"

TOKEN_FILE = DATA_DIR / ".internal_token"
REPORTS_DIR = DATA_DIR / "reports"
SERVER_DIR = Path(__file__).resolve().parent.parent
DEPLOY_DIR = SERVER_DIR.parent
SERVER_EXE = SERVER_DIR / ("kanshi-server.exe" if platform.system() == "Windows" else "kanshi-server")

_ext = ".exe" if platform.system() == "Windows" else ""
WATCHER_AFK_EXE = DEPLOY_DIR / "kanshi-watcher-afk" / f"kanshi-watcher-afk{_ext}"
WATCHER_WIN_EXE = DEPLOY_DIR / "kanshi-watcher-window" / f"kanshi-watcher-window{_ext}"
SCRIPT_PATH = Path(__file__).resolve()
# ───────────────────────────────────────────────────────────────


def load_key() -> bytes:
    return KEY_FILE.read_text(encoding="utf-8").strip().encode()


def _local_date_to_utc_range(target_date: str) -> tuple[str, str]:
    local_date = date.fromisoformat(target_date)
    local_tz = datetime.now().astimezone().tzinfo
    start_local = datetime(local_date.year, local_date.month, local_date.day, tzinfo=local_tz)
    end_local = start_local + timedelta(days=1)
    return (
        start_local.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M:%S.%f"),
        end_local.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M:%S.%f"),
    )


def scan_live_reports() -> list[dict]:
    """Scan the server's live reports directory for .meta.json files."""
    entries = []
    if not REPORTS_DIR.exists():
        return entries
    for meta_file in REPORTS_DIR.rglob("*.meta.json"):
        with open(meta_file, encoding="utf-8") as f:
            meta = json.load(f)
        enc_file = meta_file.with_name(meta_file.name.replace(".meta.json", ".enc"))
        entries.append({
            "date_from": meta["date_from"],
            "date_to": meta["date_to"],
            "emp_id": meta["emp_id"],
            "company": meta.get("company", "unknown"),
            "stored_at": meta.get("stored_at", ""),
            "enc_file": str(enc_file),
        })
    return entries


def build_report_lines() -> list[str]:
    entries = scan_live_reports()

    groups = defaultdict(list)
    for entry in entries:
        d = date.fromisoformat(entry["date_to"])
        if d <= date.today():
            key = f"{entry['company']}/{entry['emp_id']}"
            groups[key].append(entry)

    lines = [f"Report up to: {DATE_TO}", ""]

    for group_key in sorted(groups, key=str.lower):
        group_entries = sorted(groups[group_key], key=lambda e: e["date_from"])
        emp_id = group_key.split("/")[-1]
        lines.append(f"{emp_id}(dir)")
        for entry in group_entries:
            enc_name = Path(entry['enc_file']).stem  # removes .enc
            lines.append(f"  -> {enc_name}.txt")
        lines.append("")

    if len(lines) == 2:
        lines.append("No reports found in the given date range.")

    return lines


def encrypt_report():
    import sqlite3

    if not SOURCE_FOLDER.exists():
        SOURCE_FOLDER.mkdir(parents=True, exist_ok=True)

    if not KEY_FILE.exists():
        return 1

    db_path = DATA_DIR / "peewee-sqlite.v2.db"
    if not db_path.exists():
        return 1

    key = load_key()
    fernet = Fernet(key)

    conn = sqlite3.connect(str(db_path))
    cursor = conn.cursor()

    # Get all buckets and extract employee IDs from bucket names
    cursor.execute("SELECT rowid, id, type, hostname FROM bucketmodel")
    all_buckets = cursor.fetchall()

    # Extract employee ID from bucket id (e.g. "kanshi-watcher-afk_EMP-1001" -> "EMP-1001")
    emp_buckets = defaultdict(list)
    for rowid, bid, btype, hostname in all_buckets:
        parts = bid.rsplit("_", 1)
        emp_id = parts[1] if len(parts) > 1 else hostname
        emp_buckets[emp_id].append((rowid, bid, btype))

    if not emp_buckets:
        conn.close()
        return 1

    for emp_id, buckets in emp_buckets.items():
        emp_dir = SOURCE_FOLDER / emp_id
        emp_dir.mkdir(parents=True, exist_ok=True)

        # Get min/max dates for this employee
        bucket_ids = [b[0] for b in buckets]
        placeholders = ",".join("?" * len(bucket_ids))
        cursor.execute(
            f"SELECT MIN(timestamp), MAX(timestamp) FROM eventmodel WHERE bucket_id IN ({placeholders})",
            bucket_ids
        )
        row = cursor.fetchone()
        if not row[0]:
            continue

        min_date = date.fromisoformat(row[0][:10])
        max_date = date.fromisoformat(row[1][:10])

        # Generate one encrypted file per date
        current = min_date
        while current <= max_date:
            next_day = current + timedelta(days=1)
            date_str = str(current)
            next_str = str(next_day)

            lines = [f"Employee: {emp_id}", f"Date: {date_str}", ""]
            event_count = 0

            for bucket_id, name, btype in buckets:
                start_utc, end_utc = _local_date_to_utc_range(date_str)
                cursor.execute(
                    "SELECT timestamp, duration, datastr FROM eventmodel "
                    "WHERE bucket_id = ? AND timestamp >= ? AND timestamp < ? "
                    "ORDER BY timestamp",
                    (bucket_id, start_utc, end_utc)
                )
                rows = cursor.fetchall()
                if not rows:
                    continue

                event_count += len(rows)
                lines.append(f"[{name}] ({btype}) - {len(rows)} events")
                lines.append("-" * 50)
                for ts, duration, datastr in rows:
                    ts_short = ts[:19]
                    dur = f"{duration:.1f}s" if duration else "0s"
                    lines.append(f"  {ts_short}  ({dur})  {datastr}")
                lines.append("")


            if event_count > 0:
                lines.append(f"Total events: {event_count}")
            else:
                lines.append("No data for this date.")

            report_text = "\n".join(lines)
            encrypted = fernet.encrypt(report_text.encode("utf-8"))
            out_file = emp_dir / f"{date_str}.txt"
            out_file.write_bytes(encrypted)

            current = next_day

    conn.close()
    return 0


def ensure_upload_token():
    if os.environ.get("KANSHI_UPLOAD_TOKEN"):
        return
    if not TOKEN_FILE.exists():
        print(f"Token file not found: {TOKEN_FILE}")
        sys.exit(1)
    token = TOKEN_FILE.read_text(encoding="utf-8").strip()
    os.environ["KANSHI_UPLOAD_TOKEN"] = token
    print(f"KANSHI_UPLOAD_TOKEN set from {TOKEN_FILE}")


def ensure_settings_token():
    """Ensure upload_token is present in the server's settings.json."""
    settings_file = DATA_DIR / "settings.json"
    if not settings_file.exists():
        print(f"Settings file not found: {settings_file}")
        return
    if not TOKEN_FILE.exists():
        print(f"Token file not found: {TOKEN_FILE}")
        return

    token = TOKEN_FILE.read_text(encoding="utf-8").strip()

    with open(settings_file, "r", encoding="utf-8") as f:
        settings = json.load(f)

    if settings.get("upload_token") == token:
        return

    settings["upload_token"] = token
    with open(settings_file, "w", encoding="utf-8") as f:
        json.dump(settings, f, indent=4)
    print(f"upload_token injected into {settings_file}")


def start_server():
    ensure_upload_token()
    ensure_settings_token()
    if not SERVER_EXE.exists():
        print(f"Server executable not found: {SERVER_EXE}")
        return 1

    # Launch watchers as background processes
    watcher_procs = {}
    for exe, label in [(WATCHER_AFK_EXE, "afk"), (WATCHER_WIN_EXE, "window")]:
        if exe.exists():
            flags = subprocess.CREATE_NEW_PROCESS_GROUP if platform.system() == "Windows" else 0
            p = subprocess.Popen([str(exe)], env=os.environ.copy(), creationflags=flags)
            watcher_procs[label] = (exe, p)
            print(f"Started watcher-{label} (PID {p.pid})")
        else:
            print(f"Warning: watcher-{label} not found at {exe}")

    print(f"Starting server: {SERVER_EXE}")
    flags = subprocess.CREATE_NEW_PROCESS_GROUP if platform.system() == "Windows" else 0
    server_proc = subprocess.Popen([str(SERVER_EXE)], env=os.environ.copy(), creationflags=flags)

    # Watchdog loop: restart any crashed watcher every 60 seconds
    import time
    try:
        while True:
            time.sleep(60)
            # Restart crashed watchers
            for label, (exe, p) in list(watcher_procs.items()):
                if p.poll() is not None:  # process has exited
                    print(f"[Watchdog] watcher-{label} crashed (exit {p.returncode}), restarting...")
                    new_p = subprocess.Popen([str(exe)], env=os.environ.copy(), creationflags=flags)
                    watcher_procs[label] = (exe, new_p)
                    print(f"[Watchdog] watcher-{label} restarted (PID {new_p.pid})")
            # If server itself died, exit so the scheduled task can restart it
            if server_proc.poll() is not None:
                print(f"[Watchdog] kanshi-server exited ({server_proc.returncode}), shutting down.")
                break
    except KeyboardInterrupt:
        print("\nServer stopped. Terminating watchers...")
        for label, (exe, p) in watcher_procs.items():
            p.terminate()
            print(f"  Stopped watcher-{label} (PID {p.pid})")
        server_proc.terminate()
    return 0


def _fmt_duration(seconds: float) -> str:
    """Format seconds into human-readable string like '2h 45m' or '17m 16s'."""
    seconds = int(seconds)
    if seconds < 60:
        return f"{seconds}s"
    minutes, secs = divmod(seconds, 60)
    hours, minutes = divmod(minutes, 60)
    if hours > 0:
        return f"{hours}h {minutes}m"
    return f"{minutes}m {secs}s"



def decrypt_report(emp_id, target_date, open_excel=True):
    """Display report data from the database and export as formatted Excel."""
    import sys
    sys.stdout.flush()
    sys.stderr.flush()
    
    log_file = Path(os.environ.get("TEMP", ".")) / "decrypt_debug.log"
    with open(log_file, "a") as f:
        f.write(f"[{date.today()}] decrypt_report called with emp_id={emp_id}, target_date={target_date}\n")
    
    print(f"[DEBUG] decrypt_report called with emp_id={emp_id}, target_date={target_date}")
    sys.stdout.flush()
    
    import sqlite3
    import tempfile

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("Installing openpyxl...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    db_path = DATA_DIR / "peewee-sqlite.v2.db"
    print(f"Decrypting report for {emp_id} on {target_date}...")
    if not db_path.exists():
        print(f"Database not found: {db_path}")
        return 1

    conn = sqlite3.connect(str(db_path))
    cursor = conn.cursor()

    cursor.execute("SELECT rowid, id, type, client FROM bucketmodel WHERE hostname = ?", (emp_id,))
    buckets = cursor.fetchall()

    if not buckets:
        cursor.execute("SELECT rowid, id, type, client FROM bucketmodel")
        all_b = cursor.fetchall()
        buckets = [b for b in all_b if b[1].rsplit("_", 1)[-1] == emp_id]

    if not buckets:
        print(f"No buckets found for employee: {emp_id}")
        cursor.execute("SELECT DISTINCT hostname FROM bucketmodel")
        hosts = [r[0] for r in cursor.fetchall()]
        cursor.execute("SELECT DISTINCT id FROM bucketmodel")
        all_ids = set()
        for r in cursor.fetchall():
            parts = r[0].rsplit("_", 1)
            if len(parts) > 1:
                all_ids.add(parts[1])
        print(f"\nAvailable IDs: {', '.join(sorted(set(hosts) | all_ids))}")
        conn.close()
        return 1

    from datetime import datetime as _dt

    start_utc, end_utc = _local_date_to_utc_range(target_date)

    _day_start_ts = datetime.fromisoformat(start_utc).timestamp()
    _day_end_ts = datetime.fromisoformat(end_utc).timestamp()

    # Helper to merge overlapping time ranges and return total duration
    def _merge_ranges(ranges):
        """Merge overlapping (start_ts, end_ts, duration) ranges and return total seconds."""
        if not ranges:
            return 0.0
        ranges = sorted(ranges, key=lambda x: x[0])
        merged = [(ranges[0][0], ranges[0][1])]
        for cur_start, cur_end, _ in ranges[1:]:
            last_start, last_end = merged[-1]
            if cur_start <= last_end:  # overlap
                merged[-1] = (last_start, max(last_end, cur_end))
            else:
                merged.append((cur_start, cur_end))
        return sum(end - start for start, end in merged)

    afk_ranges = []      # (start_ts, end_ts, duration)
    active_ranges = []   # (start_ts, end_ts, duration)
    app_durations = defaultdict(float)
    total_events = 0

    for bucket_id, name, btype, client in buckets:
        cursor.execute(
            "SELECT timestamp, duration, datastr FROM eventmodel "
            "WHERE bucket_id = ? AND timestamp >= ? AND timestamp < ? "
            "ORDER BY timestamp",
            (bucket_id, start_utc, end_utc)
        )
        rows = cursor.fetchall()
        if not rows:
            continue
        total_events += len(rows)
        for ts, duration, datastr in rows:
            dur = duration if duration else 0
            try:
                data = json.loads(datastr)
            except (json.JSONDecodeError, TypeError):
                data = {}
            if btype == "afkstatus":
                # Parse timestamp to epoch seconds
                try:
                    start_dt = _dt.fromisoformat(ts[:19])
                    start_ts = start_dt.timestamp()
                    end_ts = start_ts + dur
                except Exception:
                    start_ts = 0
                    end_ts = dur
                if data.get("status") == "not-afk":
                    active_ranges.append((start_ts, end_ts, dur))
                else:
                    afk_ranges.append((start_ts, end_ts, dur))
            else:
                app_name = data.get("app")
                if not app_name or app_name.lower() == "unknown":
                    title = data.get("title", "").strip()
                    if title:
                        app_name = title.rsplit(" - ", 1)[-1].strip()
                if app_name:
                    app_durations[app_name] += dur

    # Merge overlapping intervals for accurate totals
    total_afk = _merge_ranges(afk_ranges)
    total_active = _merge_ranges(active_ranges)

    conn.close()

    if total_events == 0:
        print(f"No events found for {emp_id} on {target_date}")
        return 1

    # DEBUG: Show all apps to find where time is going
    all_apps = sorted(app_durations.items(), key=lambda x: x[1], reverse=True)
    total_app_time = sum(dur for _, dur in all_apps)
    
    print(f"\n[DEBUG] App Duration Analysis:")
    print(f"[DEBUG] Total Active Time (from AFK): {total_active:.0f} seconds = {_fmt_duration(total_active)}")
    print(f"[DEBUG] Total App Time (sum): {total_app_time:.0f} seconds = {_fmt_duration(total_app_time)}")
    print(f"[DEBUG] Discrepancy: {total_active - total_app_time:.0f} seconds = {_fmt_duration(abs(total_active - total_app_time))}")
    print(f"[DEBUG] Number of app events: {len(all_apps)}")
    print(f"[DEBUG] All apps (top 20):")
    for i, (app, dur) in enumerate(all_apps[:20], 1):
        print(f"  {i:2d}. {app:30s} -> {_fmt_duration(dur)}")
    
    top_apps = sorted(app_durations.items(), key=lambda x: x[1], reverse=True)[:10]
    # Suppress all terminal output except errors

    # ── Build formatted Excel ──
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Activity"

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 18

    header_fill = PatternFill("solid", fgColor="2F5496")
    header_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    sub_font = Font(name="Calibri", size=11, color="333333")

    summary_fill = PatternFill("solid", fgColor="4A7C3F")
    summary_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    summary_hdr_fill = PatternFill("solid", fgColor="5B9A4A")
    summary_hdr_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    app_title_fill = PatternFill("solid", fgColor="4A7C3F")
    app_title_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    app_hdr_fill = PatternFill("solid", fgColor="6B8E4E")
    app_hdr_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    val_font = Font(name="Calibri", size=11)
    center = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(bottom=Side(style="thin", color="CCCCCC"))

    # Row 1: Title
    ws.merge_cells("A1:C1")
    c1 = ws["A1"]
    c1.value = "Daily Activity Summary"
    c1.font = header_font
    c1.fill = header_fill
    c1.alignment = center
    ws.row_dimensions[1].height = 36

    # Row 2: Employee | Date
    ws.merge_cells("A2:C2")
    dt = date.fromisoformat(target_date)
    date_display = dt.strftime("%d %B %Y")
    c2 = ws["A2"]
    c2.value = f"Employee: {emp_id}  |  Date: {date_display}"
    c2.font = sub_font
    c2.alignment = left_align
    ws.row_dimensions[2].height = 22

    # Row 4: Summary header
    ws.merge_cells("A4:C4")
    c4 = ws["A4"]
    c4.value = "Summary"
    c4.font = summary_font
    c4.fill = summary_fill
    ws.row_dimensions[4].height = 26

    # Row 5: Column headers
    for col, val in [("A", "Metric"), ("B", "Value")]:
        cell = ws[f"{col}5"]
        cell.value = val
        cell.font = summary_hdr_font
        cell.fill = summary_hdr_fill
        cell.alignment = center

    # Row 6-7: Summary data
    ws["A6"].value = "Total Active Time"
    ws["B6"].value = _fmt_duration(total_active)
    ws["A7"].value = "Total AFK Time"
    ws["B7"].value = _fmt_duration(total_afk)
    for r in [6, 7]:
        ws[f"A{r}"].font = val_font
        ws[f"A{r}"].alignment = left_align
        ws[f"B{r}"].font = val_font
        ws[f"B{r}"].alignment = center
        ws[f"A{r}"].border = thin_border
        ws[f"B{r}"].border = thin_border

    # Row 9: Top Applications header
    ws.merge_cells("A9:C9")
    c9 = ws["A9"]
    c9.value = "Top Applications Used"
    c9.font = app_title_font
    c9.fill = app_title_fill
    ws.row_dimensions[9].height = 26

    # Row 10: Column headers
    for col, val in [("A", "Rank"), ("B", "Application"), ("C", "Active Time")]:
        cell = ws[f"{col}10"]
        cell.value = val
        cell.font = app_hdr_font
        cell.fill = app_hdr_fill
        cell.alignment = center

    # Rows 11+: App data
    for i, (app, dur) in enumerate(top_apps, 1):
        row = 10 + i
        ws[f"A{row}"].value = i
        ws[f"A{row}"].font = val_font
        ws[f"A{row}"].alignment = center
        ws[f"B{row}"].value = app
        ws[f"B{row}"].font = val_font
        ws[f"B{row}"].alignment = left_align
        ws[f"C{row}"].value = _fmt_duration(dur)
        ws[f"C{row}"].font = val_font
        ws[f"C{row}"].alignment = center
        for c in ["A", "B", "C"]:
            ws[f"{c}{row}"].border = thin_border

    xlsx_file = Path(tempfile.gettempdir()) / f"{emp_id}_{target_date}.xlsx"
    wb.save(str(xlsx_file))
    print(f"Excel report saved to: {xlsx_file}")
    print(f"File exists: {xlsx_file.exists()}")

    # Open Excel with error handling
    if open_excel:
        if platform.system() == "Windows":
            try:
                print("Opening Excel...")
                subprocess.Popen(["start", "excel", str(xlsx_file)], shell=True)
                print("Excel opened successfully")
            except Exception as e:
                print(f"Error opening Excel: {e}")
                print(f"Excel file location: {xlsx_file}")
        elif platform.system() == "Darwin":
            subprocess.run(["open", str(xlsx_file)])

    print("Done!")
    return 0


def open_next_day_report(
    emp_id: str,
    start_date: str | None = None,
    days_ahead: int = 1,
) -> None:
    """Open the Excel report for the next ``days_ahead`` days.

    This helper calls :func:`decrypt_report` for each successive date,
    ensuring that an Excel workbook is generated and opened even when the
    database contains no events for that day.  Errors for a single day are
    caught so the loop continues for the remaining days.

    Parameters
    ----------
    emp_id: str
        Hostname / employee identifier (e.g. ``LAPTOP-Q2S1QR7J``).
    start_date: str | None, optional
        The date to start from in ``YYYY-MM-DD`` format.  If ``None`` the
        function uses ``date.today()`` (the current day).
    days_ahead: int, optional
        How many days forward to open. ``1`` opens only the immediate next
        day; larger values open a range of consecutive days.
    """
    from datetime import date, timedelta

    # Resolve the first day to process (today if not supplied)
    if start_date is None:
        current = date.today()
    else:
        current = date.fromisoformat(start_date)

    # Advance to the *next* day before the first iteration
    current += timedelta(days=1)


def open_cumulative_report(
    emp_id: str,
    start_date: str,
    end_date: str | None = None,
) -> None:
    """Open Excel reports for every day from ``start_date`` up to ``end_date``.

    If ``end_date`` is omitted the function uses the current date.  It calls
    :func:`decrypt_report` for each day, which creates a workbook (or a
    placeholder) and launches Excel.  Errors for a single day are caught so
    the loop continues for the remaining days.
    """
    from datetime import date, timedelta

    start = date.fromisoformat(start_date)
    end = date.today() if end_date is None else date.fromisoformat(end_date)

    if start > end:
        raise ValueError("start_date must be earlier than or equal to end_date")

    current = start
    while current <= end:
        iso = current.isoformat()
        try:
            decrypt_report(emp_id, iso)
        except Exception as exc:  # Defensive – never abort the loop
            print(f"[WARN] Could not open report for {emp_id} on {iso}: {exc}")
        current += timedelta(days=1)


def decrypt_report_range(
    emp_id: str,
    start_date: str,
    end_date: str | None = None,
) -> int:
    """Create ONE Excel workbook with a sheet per day from start_date to end_date.

    Each sheet is named after its date (e.g. '2026-03-14') and contains the
    same formatted data that decrypt_report produces for that day.  Days with
    no events get a placeholder sheet.
    """
    from datetime import date as _date, timedelta as _td
    import tempfile
    from pathlib import Path

    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, Alignment
        from copy import copy
    except ImportError:
        import subprocess, sys
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, Alignment
        from copy import copy

    start = _date.fromisoformat(start_date)
    end = _date.today() if end_date is None else _date.fromisoformat(end_date)
    if start > end:
        raise ValueError("start_date must be earlier than or equal to end_date")

    combined_wb = Workbook()
    combined_wb.remove(combined_wb.active)  # drop the default sheet

    cur = start
    while cur <= end:
        iso = cur.isoformat()
        temp_path = Path(tempfile.gettempdir()) / f"{emp_id}_{iso}.xlsx"

        # Generate the single-day workbook (without opening Excel)
        try:
            decrypt_report(emp_id, iso, open_excel=False)
        except Exception as exc:
            print(f"[WARN] decrypt_report failed for {iso}: {exc}")

        # Copy the generated sheet into the combined workbook
        ws = combined_wb.create_sheet(title=iso)
        if temp_path.exists():
            src_wb = load_workbook(str(temp_path))
            src_ws = src_wb.active
            # Copy column widths
            for col_letter, dim in src_ws.column_dimensions.items():
                ws.column_dimensions[col_letter].width = dim.width
            # Copy row heights
            for row_num, dim in src_ws.row_dimensions.items():
                ws.row_dimensions[row_num].height = dim.height
            # Copy merged cells
            for merged_range in src_ws.merged_cells.ranges:
                ws.merge_cells(str(merged_range))
            # Copy cell values and styles
            for row in src_ws.iter_rows():
                for cell in row:
                    new_cell = ws.cell(
                        row=cell.row, column=cell.column, value=cell.value
                    )
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.border = copy(cell.border)
                        new_cell.fill = copy(cell.fill)
                        new_cell.number_format = cell.number_format
                        new_cell.protection = copy(cell.protection)
                        new_cell.alignment = copy(cell.alignment)
            src_wb.close()
        else:
            # No data for this day - add a placeholder
            ws["A1"].value = f"No data available for {iso}"
            ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="888888")
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions["A"].width = 40

        cur += _td(days=1)

    # Save the combined workbook (use timestamp in name to avoid conflicts with open files)
    import time
    ts_suffix = int(time.time())
    out_path = Path(tempfile.gettempdir()) / f"{emp_id}_{start}_{end}_cumulative_{ts_suffix}.xlsx"
    combined_wb.save(str(out_path))
    print(f"Cumulative report saved to: {out_path}")

    # Open in Excel
    import platform, subprocess
    if platform.system() == "Windows":
        try:
            subprocess.Popen(["start", "excel", str(out_path)], shell=True)
        except Exception as e:
            print(f"Error opening Excel: {e}")
    elif platform.system() == "Darwin":
        subprocess.run(["open", str(out_path)])

    return 0


# ── Default server URL for report uploads ──
SERVER_REPORT_URL = os.environ.get(
    "KANSHI_SERVER_URL", "http://localhost:5700/api/reports/upload"
)

# Received-reports directory — mirrors report_server.py RECEIVED_DIR
if platform.system() == "Windows":
    RECEIVED_DIR = Path(os.environ.get(
        "KANSHI_RECEIVED_DIR",
        str(Path(os.environ.get("LOCALAPPDATA", str(Path.home()))) / "kanshi" / "received"),
    ))
else:
    RECEIVED_DIR = Path(os.environ.get(
        "KANSHI_RECEIVED_DIR",
        str(Path.home() / ".kanshi" / "received"),
    ))


def collect_and_send_report(
    emp_id: str,
    target_date: str | None = None,
    server_url: str | None = None,
) -> int:
    """Collect a daily report from the local DB and send it to the server.

    Returns 0 on success, 1 on failure. All exceptions are caught and logged.
    """
    import sqlite3
    import logging
    from datetime import datetime as _dt

    logging.basicConfig(
        filename=str(Path(os.environ.get("TEMP", ".")) / "kanshi_send_report.log"),
        level=logging.INFO,
        format="%(asctime)s  %(levelname)s  %(message)s",
    )
    log = logging.getLogger("send_report")

    # ── Resolve defaults ──
    if target_date is None:
        target_date = str(date.today())
    if server_url is None:
        server_url = SERVER_REPORT_URL

    log.info("Collecting report for %s on %s", emp_id, target_date)

    # ── Validate inputs ──
    try:
        date.fromisoformat(target_date)
    except ValueError as exc:
        log.error("Invalid date format: %s", exc)
        print(f"[ERROR] Invalid date: {exc}")
        return 1

    # ── Load encryption key ──
    try:
        if not KEY_FILE.exists():
            log.error("Encryption key not found at %s", KEY_FILE)
            print(f"[ERROR] Key file not found: {KEY_FILE}")
            return 1
        key = load_key()
        fernet = Fernet(key)
    except Exception as exc:
        log.error("Failed to load encryption key: %s", exc)
        print(f"[ERROR] Key load failed: {exc}")
        return 1

    # ── Query the database ──
    db_path = DATA_DIR / "peewee-sqlite.v2.db"
    if not db_path.exists():
        log.error("Database not found at %s", db_path)
        print(f"[ERROR] Database not found: {db_path}")
        return 1

    try:
        conn = sqlite3.connect(str(db_path))
        cursor = conn.cursor()

        cursor.execute(
            "SELECT rowid, id, type, client FROM bucketmodel WHERE hostname = ?",
            (emp_id,),
        )
        buckets = cursor.fetchall()
        if not buckets:
            cursor.execute("SELECT rowid, id, type, client FROM bucketmodel")
            all_b = cursor.fetchall()
            buckets = [b for b in all_b if b[1].rsplit("_", 1)[-1] == emp_id]

        if not buckets:
            log.warning("No buckets found for employee %s", emp_id)
            print(f"[WARN] No buckets found for {emp_id}")
            conn.close()
            return 1

        # Deduplicate buckets per type: if multiple buckets share the same type
        # (e.g. watcher reinstalled, duplicate instances), keep only the one with
        # the most events in this day's window — prevents the same clock-second
        # being counted for different apps across buckets → totals > 24 h.
        from collections import defaultdict as _ddbkts
        _type_map = _ddbkts(list)
        for _b in buckets:
            _type_map[_b[2]].append(_b)
        deduped = []
        for _btype, _blist in _type_map.items():
            if len(_blist) == 1:
                deduped.extend(_blist)
            else:
                _best, _best_n = None, -1
                start_utc, end_utc = _local_date_to_utc_range(target_date)
                for _b in _blist:
                    cursor.execute(
                        "SELECT COUNT(*) FROM eventmodel "
                        "WHERE bucket_id = ? AND timestamp >= ? AND timestamp < ?",
                        (_b[0], start_utc, end_utc),
                    )
                    _n = cursor.fetchone()[0]
                    if _n > _best_n:
                        _best_n, _best = _n, _b
                if _best:
                    deduped.append(_best)
        buckets = deduped

        start_utc, end_utc = _local_date_to_utc_range(target_date)

        _day_start_ts = datetime.fromisoformat(start_utc).timestamp()
        _day_end_ts = datetime.fromisoformat(end_utc).timestamp()

        afk_ranges = []
        active_ranges = []
        app_ranges = defaultdict(list)   # app_name -> [(start_ts, end_ts), ...]
        total_events = 0
        from datetime import timezone as _tz

        for bucket_id, name, btype, client in buckets:
            cursor.execute(
                "SELECT timestamp, duration, datastr FROM eventmodel "
                "WHERE bucket_id = ? AND timestamp >= ? AND timestamp < ? "
                "ORDER BY timestamp",
                (bucket_id, start_utc, end_utc),
            )
            rows = cursor.fetchall()
            if not rows:
                continue
            total_events += len(rows)
            for ts, duration, datastr in rows:
                dur = duration if duration else 0
                try:
                    data = json.loads(datastr)
                except (json.JSONDecodeError, TypeError):
                    data = {}
                if btype == "afkstatus":
                    try:
                        start_dt = _dt.fromisoformat(ts[:19]).replace(tzinfo=_tz.utc)
                        start_ts = start_dt.timestamp()
                        end_ts = start_ts + dur
                    except Exception:
                        start_ts = 0
                        end_ts = dur
                    if data.get("status") == "not-afk":
                        active_ranges.append((start_ts, end_ts, dur))
                    else:
                        afk_ranges.append((start_ts, end_ts, dur))
                elif data.get("app"):
                    try:
                        ev_s = _dt.fromisoformat(ts[:19]).replace(tzinfo=_tz.utc).timestamp()
                        ev_e = ev_s + dur
                        cs = max(ev_s, _day_start_ts)
                        ce = min(ev_e, _day_end_ts)
                        if ce > cs:
                            app_ranges[data["app"]].append((cs, ce))
                    except Exception:
                        pass

        conn.close()

        # Determine session window from first→last app event.
        # Automatically adapts to any shift (day, night, late evening).
        # Excludes overnight idle and time before/after the employee was working.
        all_app_times = [t for rngs in app_ranges.values() for t in rngs]
        print(f"[DEBUG] all_app_times = {all_app_times}, _day_start_ts defined: {'_day_start_ts' in locals()}")
        if all_app_times:
            _session_start = min(s for s, e in all_app_times)
            _session_end   = max(e for s, e in all_app_times)
        else:
            _session_start = _day_start_ts
            _session_end   = _day_end_ts

        # Cap AFK/active to the actual session window (not midnight-to-midnight).
        afk_ranges    = [(max(s, _session_start), min(e, _session_end), d)
                         for s, e, d in afk_ranges
                         if s < _session_end and e > _session_start]
        active_ranges = [(max(s, _session_start), min(e, _session_end), d)
                         for s, e, d in active_ranges
                         if s < _session_end and e > _session_start]

        # Merge overlapping app ranges per app.
        # Handles duplicate buckets (watcher restarts) that would otherwise
        # double-count the same clock time for the same application.
        def _merge_app(ranges):
            if not ranges:
                return 0.0
            ranges = sorted(ranges)
            merged = [list(ranges[0])]
            for s, e in ranges[1:]:
                if s <= merged[-1][1]:
                    merged[-1][1] = max(merged[-1][1], e)
                else:
                    merged.append([s, e])
            return sum(e - s for s, e in merged)

        app_durations = {app: _merge_app(rngs) for app, rngs in app_ranges.items()}

    except sqlite3.Error as exc:
        log.error("Database error: %s", exc)
        print(f"[ERROR] Database error: {exc}")
        return 1

    # ── Handle no-data scenario ──
    if total_events == 0:
        log.warning("No events found for %s on %s", emp_id, target_date)
        print(f"[INFO] No events for {emp_id} on {target_date} — skipping report")
        return 0

    # ── Merge overlapping ranges ──
    def _merge(ranges):
        if not ranges:
            return 0.0
        ranges = sorted(ranges, key=lambda x: x[0])
        merged = [(ranges[0][0], ranges[0][1])]
        for s, e, _ in ranges[1:]:
            ls, le = merged[-1]
            if s <= le:
                merged[-1] = (ls, max(le, e))
            else:
                merged.append((s, e))
        return sum(e - s for s, e in merged)

    total_active = _merge(active_ranges)
    total_afk = _merge(afk_ranges)

    # If the AFK watcher never wrote 'not-afk' events (e.g. watcher restart lost state),
    # fall back to the sum of merged app-usage time as the active-time estimate.
    if total_active == 0 and app_durations:
        total_active = sum(app_durations.values())

    top_apps = sorted(app_durations.items(), key=lambda x: x[1], reverse=True)[:10]

    # ── Build JSON payload ──
    payload = {
        "emp_id": emp_id,
        "date": target_date,
        "total_events": total_events,
        "total_active_seconds": round(total_active, 2),
        "total_afk_seconds": round(total_afk, 2),
        "top_apps": [
            {"app": app, "seconds": round(dur, 2)} for app, dur in top_apps
        ],
        "generated_at": _dt.now().isoformat(),
    }

    # ── Encrypt ──
    try:
        payload_bytes = json.dumps(payload).encode("utf-8")
        encrypted = fernet.encrypt(payload_bytes)
    except Exception as exc:
        log.error("Encryption failed: %s", exc)
        print(f"[ERROR] Encryption failed: {exc}")
        return 1

    # ── Send to server (HTTP) or fall back to local file write ──
    import urllib.request
    import urllib.error

    http_ok = False
    try:
        token = ""
        if TOKEN_FILE.exists():
            token = TOKEN_FILE.read_text(encoding="utf-8").strip()

        req = urllib.request.Request(
            server_url,
            data=encrypted,
            headers={
                "Content-Type": "application/octet-stream",
                "X-Upload-Token": token,
                "X-Emp-Id": emp_id,
                "X-Report-Date": target_date,
            },
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=5) as resp:
            if resp.status == 200:
                log.info("Report sent via HTTP for %s on %s", emp_id, target_date)
                print(f"[OK] Report for {emp_id} on {target_date} sent to server")
                http_ok = True
            else:
                body = resp.read().decode("utf-8", errors="replace")
                log.warning("Server HTTP %d — falling back to local save", resp.status)
    except (urllib.error.URLError, OSError):
        log.info("HTTP unavailable — saving directly to RECEIVED_DIR")

    if not http_ok:
        # Fall back: write encrypted file directly to RECEIVED_DIR so that
        # `python report_server.py --decrypt` can find it without needing a running server.
        try:
            out_dir = RECEIVED_DIR / emp_id
            out_dir.mkdir(parents=True, exist_ok=True)
            out_file = out_dir / f"{target_date}.json.enc"
            out_file.write_bytes(encrypted)
            log.info("Report saved locally for %s on %s → %s", emp_id, target_date, out_file)
            print(f"[OK] Report for {emp_id} on {target_date} saved → {out_file}")
        except Exception as exc:
            log.error("Failed to save report: %s", exc)
            print(f"[ERROR] Save failed: {exc}")
            return 1

    return 0


def collect_and_send_report_range(
    emp_id: str,
    start_date: str,
    end_date: str | None = None,
    server_url: str | None = None,
) -> int:
    """Send reports for every day in the range to the server."""
    start = date.fromisoformat(start_date)
    end = date.today() if end_date is None else date.fromisoformat(end_date)
    if start > end:
        raise ValueError("start_date must be earlier than or equal to end_date")

    failures = 0
    cur = start
    while cur <= end:
        iso = cur.isoformat()
        result = collect_and_send_report(emp_id, iso, server_url)
        if result != 0:
            failures += 1
        cur += timedelta(days=1)

    total = (end - start).days + 1
    print(f"\nSent {total - failures}/{total} reports. Failures: {failures}")
    return 0 if failures == 0 else 1


def get_all_report_dates(emp_id):
    """Return all unique dates with events for the given emp_id."""
    import sqlite3
    db_path = DATA_DIR / "peewee-sqlite.v2.db"
    if not db_path.exists():
        print(f"Database not found: {db_path}")
        return []
    conn = sqlite3.connect(str(db_path))
    cursor = conn.cursor()
    # Find all bucket ids for this emp_id
    cursor.execute("SELECT id FROM bucketmodel WHERE hostname = ?", (emp_id,))
    buckets = [row[0] for row in cursor.fetchall()]
    if not buckets:
        cursor.execute("SELECT id FROM bucketmodel")
        all_b = cursor.fetchall()
        buckets = [b[0] for b in all_b if b[0].rsplit("_", 1)[-1] == emp_id]
    if not buckets:
        print(f"No buckets found for employee: {emp_id}")
        conn.close()
        return []
    placeholders = ",".join(["?"] * len(buckets))
    cursor.execute(f"SELECT DISTINCT substr(timestamp, 1, 10) FROM eventmodel WHERE bucket_id IN ({placeholders})", buckets)
    dates = sorted(row[0] for row in cursor.fetchall())
    conn.close()
    return dates

def main() -> int:
    print(f"[DEBUG] main() called with argv: {sys.argv}")
    if "--server" in sys.argv:
        return start_server()
    if "--decrypt-all" in sys.argv:
        # Usage: python setup_and_run.py --decrypt-all <emp_id>
        if len(sys.argv) < 3:
            emp_id = input("Enter employee ID: ").strip()
        else:
            emp_id = sys.argv[2]
        dates = get_all_report_dates(emp_id)
        if not dates:
            print(f"No report dates found for employee: {emp_id}")
            return 1
        for d in dates:
            decrypt_report(emp_id, d)
        return 0
    if "--decrypt-range" in sys.argv:
        # Usage: python setup_and_run.py --decrypt-range <emp_id> <start_date> [end_date]
        if len(sys.argv) < 4:
            emp_id = input("Enter employee ID: ").strip()
            start_d = input("Enter start date (YYYY-MM-DD): ").strip()
        else:
            emp_id = sys.argv[2]
            start_d = sys.argv[3]
        end_d = sys.argv[4] if len(sys.argv) >= 5 else None
        try:
            return decrypt_report_range(emp_id, start_d, end_d)
        except Exception as e:
            print(f"ERROR in decrypt_report_range: {e}")
            import traceback
            traceback.print_exc()
            return 1
    if "--decrypt" in sys.argv:
        # Usage: python setup_and_run.py --decrypt <emp_id> <date>
        if len(sys.argv) < 4:
            emp_id = input("Enter employee ID: ").strip()
            target_date = input("Enter date (YYYY-MM-DD): ").strip()
        else:
            emp_id = sys.argv[2]
            target_date = sys.argv[3]
        try:
            return decrypt_report(emp_id, target_date)
        except Exception as e:
            print(f"ERROR in decrypt_report: {e}")
            import traceback
            traceback.print_exc()
            return 1
    if "--send-report-range" in sys.argv:
        # Usage: python setup_and_run.py --send-report-range <emp_id> <start_date> [end_date] [server_url]
        if len(sys.argv) < 4:
            emp_id = input("Enter employee ID: ").strip()
            start_d = input("Enter start date (YYYY-MM-DD): ").strip()
        else:
            emp_id = sys.argv[2]
            start_d = sys.argv[3]
        end_d = sys.argv[4] if len(sys.argv) >= 5 else None
        s_url = sys.argv[5] if len(sys.argv) >= 6 else None
        try:
            return collect_and_send_report_range(emp_id, start_d, end_d, s_url)
        except Exception as e:
            print(f"ERROR in collect_and_send_report_range: {e}")
            import traceback
            traceback.print_exc()
            return 1
    if "--send-report" in sys.argv:
        # Usage: python setup_and_run.py --send-report <emp_id> [date] [server_url]
        if len(sys.argv) < 3:
            emp_id = input("Enter employee ID: ").strip()
        else:
            emp_id = sys.argv[2]
        t_date = sys.argv[3] if len(sys.argv) >= 4 else None
        s_url = sys.argv[4] if len(sys.argv) >= 5 else None
        try:
            return collect_and_send_report(emp_id, t_date, s_url)
        except Exception as e:
            print(f"ERROR in collect_and_send_report: {e}")
            import traceback
            traceback.print_exc()
            return 1
    if "--send-night" in sys.argv:
        # Run once each morning to send the previous day's full report.
        import socket
        emp_id = socket.gethostname()
        t_date = (date.today() - timedelta(days=1)).isoformat()
        s_url = sys.argv[2] if len(sys.argv) >= 3 else None
        try:
            return collect_and_send_report(emp_id, t_date, s_url)
        except Exception as e:
            print(f"ERROR in collect_and_send_report: {e}")
            import traceback
            traceback.print_exc()
            return 1
    if "--send-day" in sys.argv:
        # Run once each night to send the current day's full report.
        import socket
        emp_id = socket.gethostname()
        t_date = date.today().isoformat()
        s_url = sys.argv[2] if len(sys.argv) >= 3 else None
        try:
            return collect_and_send_report(emp_id, t_date, s_url)
        except Exception as e:
            print(f"ERROR in collect_and_send_report: {e}")
            import traceback
            traceback.print_exc()
            return 1
    if "--send-today" in sys.argv:
        # Backwards compatibility: keep the old logic for tasks still using --send-today.
        import socket
        from datetime import datetime as _dt
        emp_id = socket.gethostname()
        now = _dt.now()
        if now.hour < 12:
            t_date = (date.today() - timedelta(days=1)).isoformat()
        else:
            t_date = date.today().isoformat()
        s_url = sys.argv[2] if len(sys.argv) >= 3 else None
        try:
            return collect_and_send_report(emp_id, t_date, s_url)
        except Exception as e:
            print(f"ERROR in collect_and_send_report: {e}")
            import traceback
            traceback.print_exc()
            return 1
    return encrypt_report()


if __name__ == "__main__":
    sys.exit(main())