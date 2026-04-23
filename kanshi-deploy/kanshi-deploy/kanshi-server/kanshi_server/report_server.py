#!/usr/bin/env python3
"""
report_server.py — Kanshi encrypted report receiver + Excel generator

USAGE
─────
# On the SERVER — start the receiver once:
python report_server.py --serve
python report_server.py --serve --host 0.0.0.0 --port 5700

# On the SERVER — decrypt received reports and save to Excel:
python report_server.py --decrypt LAPTOP-Q2S1QR7J 2026-04-01 2026-04-20
python report_server.py --decrypt-all 2026-04-01 2026-04-20

# On the SERVER — list all received reports:
python report_server.py --list

# On the CLIENT — send reports to the server:
python setup_and_run.py --send-report LAPTOP-Q2S1QR7J 2026-04-15
python setup_and_run.py --send-report-range LAPTOP-Q2S1QR7J 2026-04-01 2026-04-20
"""
from __future__ import annotations

import json
import os
import platform
import sys
from datetime import date, datetime, timedelta
from pathlib import Path


# ── Paths ─────────────────────────────────────────────────────────────────────

def _get_documents_dir() -> Path:
    if platform.system() == "Windows":
        onedrive = os.environ.get("OneDrive", "")
        if onedrive:
            od = Path(onedrive) / "Documents"
            if od.exists():
                return od
    return Path.home() / "Documents"


SOURCE_FOLDER = _get_documents_dir() / "KanshiReports"
KEY_FILE      = SOURCE_FOLDER / ".key"

if platform.system() == "Windows":
    DATA_DIR = Path(os.environ.get("LOCALAPPDATA", "")) / "kanshi" / "kanshi" / "kanshi-server"
elif platform.system() == "Darwin":
    DATA_DIR = Path.home() / "Library" / "Application Support" / "kanshi" / "kanshi-server"
else:
    DATA_DIR = Path.home() / ".local" / "share" / "kanshi" / "kanshi-server"

TOKEN_FILE = DATA_DIR / ".internal_token"

# Encrypted reports land here after upload
if platform.system() == "Windows":
    _default_received = Path(os.environ.get("LOCALAPPDATA", Path.home())) / "kanshi" / "received"
else:
    _default_received = Path.home() / ".kanshi" / "received"

RECEIVED_DIR = Path(os.environ.get("KANSHI_RECEIVED_DIR", str(_default_received)))

# Excel files are saved here
OUTPUT_DIR = Path(os.environ.get("KANSHI_OUTPUT_DIR", str(_get_documents_dir() / "KanshiExcelReports")))


# ── Helpers ───────────────────────────────────────────────────────────────────

def _load_key() -> bytes:
    if not KEY_FILE.exists():
        raise FileNotFoundError(f"Encryption key not found: {KEY_FILE}")
    return KEY_FILE.read_text(encoding="utf-8").strip().encode()


def _check_token(provided: str) -> bool:
    if not TOKEN_FILE.exists():
        return True  # no token file → open / local mode
    return TOKEN_FILE.read_text(encoding="utf-8").strip() == provided


def _decrypt_payload(path: Path) -> dict:
    from cryptography.fernet import Fernet
    fernet = Fernet(_load_key())
    raw    = fernet.decrypt(path.read_bytes()).decode("utf-8")
    return json.loads(raw)


def _fmt(seconds: float) -> str:
    s = int(seconds)
    h, r = divmod(s, 3600)
    m, s = divmod(r, 60)
    if h:
        return f"{h}h {m}m {s}s"
    if m:
        return f"{m}m {s}s"
    return f"{s}s"


# ── Excel builder ─────────────────────────────────────────────────────────────

def _build_excel(emp_id: str, reports: list[dict]) -> Path:
    try:
        from copy import copy
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError:
        print("[ERROR] openpyxl is required:  pip install openpyxl")
        sys.exit(1)

    DARK   = "1A1A2E"
    ACCENT = "E94560"
    MID    = "16213E"
    LGREY  = "F5F6FA"
    WHITE  = "FFFFFF"
    GREEN  = "27AE60"
    RED    = "E74C3C"

    def _fill(c):
        return PatternFill("solid", fgColor=c)

    def _font(bold=False, color=None, size=11):
        return Font(bold=bold, color=color or "000000", size=size)

    def _border():
        t = Side(style="thin", color="DDDDDD")
        return Border(left=t, right=t, top=t, bottom=t)

    def _center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    wb = Workbook()
    wb.remove(wb.active)

    for rpt in sorted(reports, key=lambda x: x["date"]):
        if (not rpt.get("top_apps")
                and rpt.get("total_active_seconds", 0) == 0
                and rpt.get("total_afk_seconds", 0) == 0):
            continue
        ws   = wb.create_sheet(title=rpt["date"][:10])
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 16
        row = 1

        # ── Title ──
        ws.merge_cells(f"A{row}:D{row}")
        tc = ws[f"A{row}"]
        tc.value     = f"Kanshi Activity Report  —  {emp_id}  —  {rpt['date'][:10]}"
        tc.fill      = _fill(DARK)
        tc.font      = _font(bold=True, color=WHITE, size=13)
        tc.alignment = _center()
        ws.row_dimensions[row].height = 32
        row += 2

        # ── Summary boxes ──
        summary = [
            ("Active Time",  _fmt(rpt.get("total_active_seconds", 0)), GREEN),
            ("AFK Time",     _fmt(rpt.get("total_afk_seconds",   0)),  RED),
        ]
        for col, (lbl, val, color) in enumerate(summary, start=1):
            hc = ws.cell(row=row,     column=col, value=lbl)
            vc = ws.cell(row=row + 1, column=col, value=val)
            hc.fill      = _fill(MID)
            hc.font      = _font(bold=True, color=WHITE, size=10)
            hc.alignment = _center()
            hc.border    = _border()
            vc.font      = _font(bold=True, color=color, size=14)
            vc.alignment = _center()
            vc.border    = _border()
            vc.fill      = _fill(LGREY)
            ws.row_dimensions[row].height     = 18
            ws.row_dimensions[row + 1].height = 30
        row += 3

        # ── Top apps table ──
        apps = rpt.get("top_apps", [])
        if apps:
            for col, hdr in enumerate(["Application", "Duration"], start=1):
                c = ws.cell(row=row, column=col, value=hdr)
                c.fill      = _fill(ACCENT)
                c.font      = _font(bold=True, color=WHITE, size=10)
                c.alignment = _center()
                c.border    = _border()
            ws.row_dimensions[row].height = 20
            row += 1

            for i, app in enumerate(apps):
                bg  = LGREY if i % 2 == 0 else WHITE
                for col, val in enumerate([app["app"], _fmt(app["seconds"])], start=1):
                    c            = ws.cell(row=row, column=col, value=val)
                    c.fill       = _fill(bg)
                    c.font       = _font(size=10)
                    c.alignment  = Alignment(
                        horizontal="left" if col == 1 else "center",
                        vertical="center",
                    )
                    c.border     = _border()
                ws.row_dimensions[row].height = 17
                row += 1
        else:
            c      = ws.cell(row=row, column=1, value="No application data recorded.")
            c.font = _font(color="AAAAAA")
            row   += 1

        # ── Footer ──
        row += 1
        gen = rpt.get("generated_at", "")[:19]
        ws.cell(row=row, column=1, value=f"Generated: {gen}").font = _font(color="AAAAAA", size=9)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    dates   = sorted(r["date"][:10] for r in reports)
    tag     = f"{dates[0]}_to_{dates[-1]}" if len(dates) > 1 else dates[0]
    out     = OUTPUT_DIR / f"KanshiReport_{emp_id}_{tag}.xlsx"
    if out.exists():
        ts  = datetime.now().strftime("%H%M%S")
        out = OUTPUT_DIR / f"KanshiReport_{emp_id}_{tag}_{ts}.xlsx"
    wb.save(str(out))
    return out


# ── CLI commands ──────────────────────────────────────────────────────────────

def cmd_list() -> int:
    if not RECEIVED_DIR.exists():
        print(f"[INFO] No received reports yet. Directory: {RECEIVED_DIR}")
        return 0
    found = False
    for emp_dir in sorted(RECEIVED_DIR.iterdir()):
        if not emp_dir.is_dir():
            continue
        files = sorted(
            f.stem.replace(".json", "") for f in emp_dir.glob("*.json.enc")
        )
        if files:
            print(f"  {emp_dir.name}:  {files[0]}  →  {files[-1]}  ({len(files)} days)")
            found = True
    if not found:
        print("[INFO] No reports found.")
    return 0


def cmd_decrypt(emp_ids: list[str], start_date: str, end_date: str | None) -> int:
    start = date.fromisoformat(start_date)
    end   = date.today() if not end_date else date.fromisoformat(end_date)

    if not RECEIVED_DIR.exists():
        print(f"[ERROR] Received-reports directory not found: {RECEIVED_DIR}")
        return 1

    overall_rc = 0
    for emp_id in emp_ids:
        emp_dir = RECEIVED_DIR / emp_id
        if not emp_dir.exists():
            print(f"[WARN] No reports directory for {emp_id}")
            continue

        reports: list[dict] = []
        cur = start
        while cur <= end:
            fpath = emp_dir / f"{cur.isoformat()}.json.enc"
            if fpath.exists():
                try:
                    reports.append(_decrypt_payload(fpath))
                    print(f"  [OK] Decrypted {emp_id}/{cur.isoformat()}")
                except Exception as exc:
                    print(f"  [WARN] {fpath.name}: {exc}")
            cur += timedelta(days=1)

        if not reports:
            print(f"[WARN] No decryptable reports for {emp_id} in {start_date} → {end_date or 'today'}")
            overall_rc = 1
            continue

        try:
            out = _build_excel(emp_id, reports)
            print(f"\n[OK] Excel saved → {out}")
            if platform.system() == "Windows":
                os.startfile(str(out))
            elif platform.system() == "Darwin":
                import subprocess
                subprocess.run(["open", str(out)])
        except Exception as exc:
            import traceback
            print(f"[ERROR] Excel generation failed for {emp_id}: {exc}")
            traceback.print_exc()
            overall_rc = 1

    return overall_rc


# ── Flask receiver ────────────────────────────────────────────────────────────

def cmd_serve(host: str, port: int) -> None:
    try:
        from flask import Flask, abort, jsonify, request
    except ImportError:
        print("[ERROR] Flask is required:  pip install flask")
        sys.exit(1)

    app = Flask(__name__)

    @app.route("/", methods=["GET"])
    def index():
        from flask import Response
        employee_count = 0
        report_count   = 0
        if RECEIVED_DIR.exists():
            for emp_dir in RECEIVED_DIR.iterdir():
                if emp_dir.is_dir():
                    n = len(list(emp_dir.glob("*.json.enc")))
                    if n:
                        employee_count += 1
                        report_count   += n
        html = f"""<!DOCTYPE html>
<html><head><title>Kanshi Report Receiver</title>
<style>
  body {{ font-family: Arial, sans-serif; background:#1A1A2E; color:#F5F6FA; padding:40px; }}
  h1   {{ color:#E94560; }}
  .box {{ background:#16213E; border-radius:8px; padding:20px; margin:16px 0; max-width:520px; }}
  a    {{ color:#E94560; }}
  code {{ background:#0F3460; padding:2px 6px; border-radius:4px; }}
</style></head><body>
<h1>Kanshi Report Receiver</h1>
<div class="box">
  <b>Status:</b> &#x2705; Running<br>
  <b>Employees with reports:</b> {employee_count}<br>
  <b>Total encrypted reports:</b> {report_count}<br>
  <b>Received dir:</b> <code>{RECEIVED_DIR}</code>
</div>
</body></html>"""
        return Response(html, mimetype="text/html")

    @app.route("/api/reports/upload", methods=["POST"])
    def upload_report():
        token = request.headers.get("X-Upload-Token", "")
        if not _check_token(token):
            abort(401, description="Invalid X-Upload-Token")

        emp_id      = request.headers.get("X-Emp-Id",      "").strip()
        report_date = request.headers.get("X-Report-Date", "").strip()

        if not emp_id or not report_date:
            abort(400, description="X-Emp-Id and X-Report-Date headers required")
        try:
            date.fromisoformat(report_date)
        except ValueError:
            abort(400, description=f"Invalid X-Report-Date: {report_date}")

        encrypted = request.get_data()
        if not encrypted:
            abort(400, description="Empty body")

        try:
            out_dir = RECEIVED_DIR / emp_id
            out_dir.mkdir(parents=True, exist_ok=True)
            (out_dir / f"{report_date}.json.enc").write_bytes(encrypted)
        except Exception as exc:
            abort(500, description=f"Failed to save: {exc}")

        print(f"  [RECEIVED] {emp_id} / {report_date}")
        return jsonify({"status": "ok", "emp_id": emp_id, "date": report_date}), 200

    @app.route("/api/reports/list")
    def list_api():
        token = request.headers.get("X-Download-Token", "") or request.args.get("token", "")
        if not _check_token(token):
            abort(401, description="Invalid or missing token")
        result: dict = {}
        if RECEIVED_DIR.exists():
            for emp_dir in sorted(RECEIVED_DIR.iterdir()):
                if not emp_dir.is_dir():
                    continue
                files = sorted(
                    f.stem.replace(".json", "") for f in emp_dir.glob("*.json.enc")
                )
                if files:
                    result[emp_dir.name] = {
                        "from": files[0], "to": files[-1], "count": len(files)
                    }
        return jsonify({"employees": result})

    print(f"[Kanshi Report Receiver]  http://{host}:{port}")
    print(f"  Received dir  : {RECEIVED_DIR}")
    print(f"  Excel out dir : {OUTPUT_DIR}")
    print(f"  Token file    : {TOKEN_FILE}")
    app.run(host=host, port=port, debug=False)


# ── main ──────────────────────────────────────────────────────────────────────

def main() -> int:
    args = sys.argv[1:]

    if not args or args[0] in ("-h", "--help"):
        print(__doc__)
        return 0

    if args[0] == "--serve":
        host, port = "127.0.0.1", 5700
        for i, a in enumerate(args):
            if a == "--host" and i + 1 < len(args):
                host = args[i + 1]
            if a == "--port" and i + 1 < len(args):
                port = int(args[i + 1])
        cmd_serve(host, port)
        return 0

    if args[0] == "--list":
        return cmd_list()

    if args[0] == "--decrypt":
        # --decrypt <emp_id> <start> [end]
        if len(args) < 3:
            print("Usage: report_server.py --decrypt <emp_id> <start_date> [end_date]")
            return 1
        return cmd_decrypt([args[1]], args[2], args[3] if len(args) >= 4 else None)

    if args[0] == "--decrypt-all":
        # --decrypt-all <start> [end]
        if len(args) < 2:
            print("Usage: report_server.py --decrypt-all <start_date> [end_date]")
            return 1
        if not RECEIVED_DIR.exists():
            print(f"[ERROR] {RECEIVED_DIR} not found")
            return 1
        emp_ids = [d.name for d in sorted(RECEIVED_DIR.iterdir()) if d.is_dir()]
        if not emp_ids:
            print("[INFO] No employee directories found.")
            return 0
        return cmd_decrypt(emp_ids, args[1], args[2] if len(args) >= 3 else None)

    print(f"[ERROR] Unknown command: {args[0]}")
    print(__doc__)
    return 1


if __name__ == "__main__":
    sys.exit(main())
